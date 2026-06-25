"""
Ingestion module — Parse raw XLSX dataset into RawMention objects.

Inputs:  Dataset.xlsx
Outputs: List[RawMention] + ingestion audit log
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import List, Tuple

import openpyxl

from app.config import settings
from app.models.mention import RawMention

logger = logging.getLogger(__name__)


# Column mapping: XLSX header → RawMention field
COLUMN_MAP = {
    "Date": "date_raw",
    "URL": "url",
    "Source Name": "source_name_raw",
    "Title": "title_raw",
    "Opening Text": "opening_text_raw",
    "Hit Sentence": "hit_sentence_raw",
    "Sentiment": "sentiment_raw",
    "Reach": "reach_raw",
}

# Columns we intentionally skip (they're empty in the dataset — classification is our job)
SKIP_COLUMNS = {"Driver", "Sub driver"}


def ingest_xlsx(
    filepath: Path | None = None,
    sheet_name: str | None = None,
) -> Tuple[List[RawMention], dict]:
    """
    Parse the XLSX dataset into a list of RawMention objects.

    Args:
        filepath: Path to the XLSX file. Defaults to config.
        sheet_name: Name of the sheet to read. Defaults to config.

    Returns:
        Tuple of (mentions, audit_log)
        - mentions: List of RawMention objects
        - audit_log: dict with ingestion statistics and issues
    """
    filepath = filepath or (settings.DATA_RAW_DIR / settings.DATASET_FILENAME)
    sheet_name = sheet_name or settings.DATASET_SHEET_NAME

    logger.info(f"Ingesting dataset from {filepath}, sheet: {sheet_name}")

    if not filepath.exists():
        raise FileNotFoundError(f"Dataset not found: {filepath}")

    wb = openpyxl.load_workbook(str(filepath), read_only=True)

    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
        )

    ws = wb[sheet_name]

    # Parse headers
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    logger.info(f"Headers found: {headers}")

    # Validate expected columns exist
    expected = set(COLUMN_MAP.keys())
    found = set(h for h in headers if h in expected)
    missing = expected - found
    if missing:
        logger.warning(f"Missing expected columns: {missing}")

    # Build column index → field name mapping
    col_mapping = {}
    for idx, header in enumerate(headers):
        if header in COLUMN_MAP:
            col_mapping[idx] = COLUMN_MAP[header]
        elif header in SKIP_COLUMNS:
            continue
        else:
            logger.debug(f"Unmapped column at index {idx}: {header}")

    # Parse rows
    mentions: List[RawMention] = []
    audit = {
        "total_rows": 0,
        "parsed_rows": 0,
        "skipped_rows": 0,
        "issues": [],
    }

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True),
        start=1,
    ):
        audit["total_rows"] += 1

        # Build field dict from row
        fields = {"row_number": row_idx}

        for col_idx, cell_value in enumerate(row):
            if col_idx in col_mapping:
                field_name = col_mapping[col_idx]

                # Handle reach — might be float from Excel
                if field_name == "reach_raw" and cell_value is not None:
                    try:
                        cell_value = int(cell_value)
                    except (ValueError, TypeError):
                        audit["issues"].append(
                            f"Row {row_idx}: Invalid reach value: {cell_value}"
                        )
                        cell_value = None

                # Handle empty strings as None
                if isinstance(cell_value, str) and cell_value.strip() == "":
                    cell_value = None

                fields[field_name] = cell_value

        # Must have at least a URL
        if not fields.get("url"):
            audit["issues"].append(f"Row {row_idx}: Missing URL, skipping")
            audit["skipped_rows"] += 1
            continue

        # Must have at least some text content
        has_text = any(
            fields.get(f)
            for f in ("title_raw", "opening_text_raw", "hit_sentence_raw")
        )
        if not has_text:
            audit["issues"].append(
                f"Row {row_idx}: No text content (title, opening, hit all empty), skipping"
            )
            audit["skipped_rows"] += 1
            continue

        try:
            mention = RawMention(**fields)
            mentions.append(mention)
            audit["parsed_rows"] += 1
        except Exception as e:
            audit["issues"].append(f"Row {row_idx}: Parse error: {e}")
            audit["skipped_rows"] += 1

    wb.close()

    logger.info(
        f"Ingestion complete: {audit['parsed_rows']} parsed, "
        f"{audit['skipped_rows']} skipped out of {audit['total_rows']} total"
    )

    return mentions, audit


def save_raw_mentions(mentions: List[RawMention], output_dir: Path | None = None) -> Path:
    """Save raw mentions as JSON for audit trail."""
    import json

    output_dir = output_dir or settings.DATA_PROCESSED_DIR
    output_path = output_dir / "01_raw_mentions.json"

    data = [m.model_dump(mode="json") for m in mentions]

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, default=str, ensure_ascii=False)

    logger.info(f"Saved {len(mentions)} raw mentions to {output_path}")
    return output_path


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    mentions, audit = ingest_xlsx()
    print(f"\nIngestion Audit:")
    print(f"  Total rows:   {audit['total_rows']}")
    print(f"  Parsed:       {audit['parsed_rows']}")
    print(f"  Skipped:      {audit['skipped_rows']}")
    if audit["issues"]:
        print(f"  Issues ({len(audit['issues'])}):")
        for issue in audit["issues"]:
            print(f"    - {issue}")
    save_raw_mentions(mentions)
